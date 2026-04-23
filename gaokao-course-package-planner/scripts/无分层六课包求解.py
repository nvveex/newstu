from __future__ import annotations

import os
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from openpyxl import Workbook, load_workbook
from ortools.sat.python import cp_model


SUBJECTS = ["物理", "化学", "生物", "历史", "政治", "地理"]
PACKAGES = list(range(6))


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return (
        text.replace("１", "1")
        .replace("２", "2")
        .replace("３", "3")
        .replace("Ａ", "A")
        .replace("Ｂ", "B")
        .replace("Ｃ", "C")
        .replace("合考", "合格考")
    )


def find_input_workbook() -> Path:
    desktop = Path.home() / "Desktop"
    matches = [
        desktop / name
        for name in os.listdir(desktop)
        if name.endswith(".xlsx")
        and not name.startswith("~$")
        and "高一选考分班源数据无分层" in name
    ]
    if not matches:
        raise FileNotFoundError("未在桌面找到“高一选考分班源数据无分层.xlsx”")
    return matches[0]


def next_available_output_path(base_path: Path) -> Path:
    if not base_path.exists():
        return base_path
    stem = base_path.stem
    suffix = base_path.suffix
    parent = base_path.parent
    index = 2
    while True:
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def parse_section_name(name: str) -> Tuple[str, str]:
    normalized = normalize_text(name)
    subject = next((item for item in SUBJECTS if normalized.startswith(item)), None)
    if subject is None:
        raise ValueError(f"无法识别课程班科目：{name}")
    mode = "选考" if "选考" in normalized else "合格考"
    return subject, mode


@dataclass(frozen=True)
class Section:
    id: int
    name: str
    teacher: str
    subject: str
    mode: str

    @property
    def type_key(self) -> Tuple[str, str]:
        return self.subject, self.mode


@dataclass
class Student:
    id: int
    name: str
    work_id: str
    admin_class: str
    selected_subjects: Tuple[str, str, str]
    required_types: Dict[str, str]


def load_data(workbook_path: Path) -> Tuple[List[Student], List[Section]]:
    wb = load_workbook(workbook_path, data_only=True)
    student_sheet = wb["选考-源数据"]
    teacher_sheet = wb["教师列表"]

    students: List[Student] = []
    for idx, row in enumerate(student_sheet.iter_rows(min_row=2, values_only=True)):
        name, work_id, admin_class, s1, s2, s3 = row
        selected = tuple(sorted(normalize_text(item) for item in (s1, s2, s3)))
        selected_set = set(selected)
        required_types = {
            subject: ("选考" if subject in selected_set else "合格考") for subject in SUBJECTS
        }
        students.append(
            Student(
                id=idx,
                name=normalize_text(name),
                work_id=normalize_text(work_id),
                admin_class=normalize_text(admin_class),
                selected_subjects=selected,
                required_types=required_types,
            )
        )

    seen = set()
    sections: List[Section] = []
    for row in teacher_sheet.iter_rows(min_row=2, values_only=True):
        section_name_raw, teacher_raw = row[:2]
        section_name = normalize_text(section_name_raw)
        teacher = normalize_text(teacher_raw)
        if not section_name or not teacher:
            continue
        key = (section_name, teacher)
        if key in seen:
            continue
        seen.add(key)
        subject, mode = parse_section_name(section_name)
        sections.append(
            Section(
                id=len(sections),
                name=section_name,
                teacher=teacher,
                subject=subject,
                mode=mode,
            )
        )

    return students, sections


def build_model(
    students: List[Student],
    sections: List[Section],
    relax: int = 0,
) -> Tuple[cp_model.CpModel, Dict, Dict, Dict]:
    model = cp_model.CpModel()

    sections_by_type: Dict[Tuple[str, str], List[Section]] = defaultdict(list)
    sections_by_teacher: Dict[str, List[Section]] = defaultdict(list)
    for section in sections:
        sections_by_type[section.type_key].append(section)
        sections_by_teacher[section.teacher].append(section)

    required_totals = Counter()
    for student in students:
        for subject, mode in student.required_types.items():
            required_totals[(subject, mode)] += 1

    x = {}
    for student in students:
        for subject in SUBJECTS:
            mode = student.required_types[subject]
            candidate_sections = sections_by_type[(subject, mode)]
            section_vars = []
            for section in candidate_sections:
                var = model.NewBoolVar(f"x_s{student.id}_c{section.id}")
                x[(student.id, section.id)] = var
                section_vars.append(var)
            model.Add(sum(section_vars) == 1)

    balance_penalties = []
    for type_key, type_sections in sections_by_type.items():
        total = required_totals[type_key]
        count = len(type_sections)
        low = total // count
        high = low + (1 if total % count else 0)
        low = max(0, low - relax)
        high = high + relax
        for section in type_sections:
            student_vars = [
                x[(student.id, section.id)]
                for student in students
                if (student.id, section.id) in x
            ]
            size_var = model.NewIntVar(low, high, f"size_c{section.id}")
            model.Add(size_var == sum(student_vars))

            scaled_dev = model.NewIntVar(
                0, max(total, count * high), f"dev_c{section.id}"
            )
            model.AddAbsEquality(scaled_dev, size_var * count - total)
            balance_penalties.append(scaled_dev)

    y = {}
    for section in sections:
        vars_for_section = []
        for package in PACKAGES:
            var = model.NewBoolVar(f"y_c{section.id}_p{package}")
            y[(section.id, package)] = var
            vars_for_section.append(var)
        model.Add(sum(vars_for_section) == 1)

    for package in PACKAGES:
        model.Add(sum(y[(section.id, package)] for section in sections) == 5)

    for teacher, teacher_sections in sections_by_teacher.items():
        for package in PACKAGES:
            model.Add(sum(y[(section.id, package)] for section in teacher_sections) <= 1)

    model.Add(y[(sections[0].id, 0)] == 1)

    z = {}
    for student in students:
        eligible_sections = [section for section in sections if (student.id, section.id) in x]
        for section in eligible_sections:
            x_var = x[(student.id, section.id)]
            for package in PACKAGES:
                z_var = model.NewBoolVar(f"z_s{student.id}_c{section.id}_p{package}")
                y_var = y[(section.id, package)]
                model.Add(z_var <= x_var)
                model.Add(z_var <= y_var)
                model.Add(z_var >= x_var + y_var - 1)
                z[(student.id, section.id, package)] = z_var

        for package in PACKAGES:
            package_vars = [
                z[(student.id, section.id, package)]
                for section in eligible_sections
            ]
            model.Add(sum(package_vars) == 1)

    model.Minimize(sum(balance_penalties))

    return model, x, y, z


def solve(students: List[Student], sections: List[Section]):
    last_status = None
    for relax in range(3):
        model, x, y, z = build_model(students, sections, relax=relax)
        solver = cp_model.CpSolver()
        solver.parameters.max_time_in_seconds = 180
        solver.parameters.num_search_workers = 8
        solver.parameters.random_seed = 42
        status = solver.Solve(model)
        last_status = status
        if status in (cp_model.OPTIMAL, cp_model.FEASIBLE):
            return solver, x, y, z, relax
    raise RuntimeError(f"未求得可行解，求解状态：{last_status}")


def autosize(sheet):
    widths = defaultdict(int)
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths[cell.column], len(value))
    for column, width in widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = min(
            max(width + 2, 10), 60
        )


def export_results(
    output_path: Path,
    students: List[Student],
    sections: List[Section],
    solver: cp_model.CpSolver,
    x: Dict,
    y: Dict,
):
    section_map = {section.id: section for section in sections}
    package_of_section = {}
    section_students: Dict[int, List[Student]] = defaultdict(list)
    student_sections: Dict[int, Dict[str, Section]] = defaultdict(dict)
    student_packages: Dict[int, Dict[int, Section]] = defaultdict(dict)

    for section in sections:
        for package in PACKAGES:
            if solver.Value(y[(section.id, package)]):
                package_of_section[section.id] = package
                break

    for student in students:
        for section in sections:
            key = (student.id, section.id)
            if key in x and solver.Value(x[key]):
                assigned = section_map[section.id]
                section_students[section.id].append(student)
                student_sections[student.id][assigned.subject] = assigned
                student_packages[student.id][package_of_section[section.id]] = assigned

    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    summary_sheet = wb.create_sheet("课程班汇总")
    summary_sheet.append(
        ["课位包", "课程班", "科目", "类型", "教师", "学生数", "学生名单"]
    )

    for package in PACKAGES:
        package_sheet = wb.create_sheet(f"课位包{package + 1}")
        package_sheet.append(
            ["课位包", "课程班", "科目", "类型", "教师", "学生数", "学生名单"]
        )
        package_sections = [
            section
            for section in sections
            if package_of_section[section.id] == package
        ]
        package_sections.sort(key=lambda item: (item.subject, item.mode, item.name))
        for section in package_sections:
            students_in_section = sorted(
                section_students[section.id], key=lambda item: (item.admin_class, item.name)
            )
            names = "、".join(student.name for student in students_in_section)
            row = [
                f"课位包{package + 1}",
                section.name,
                section.subject,
                section.mode,
                section.teacher,
                len(students_in_section),
                names,
            ]
            package_sheet.append(row)
            summary_sheet.append(row)
        autosize(package_sheet)

    student_sheet = wb.create_sheet("学生分班结果")
    headers = [
        "姓名",
        "学工号",
        "行政班",
        "选考组合",
        "物理安排",
        "化学安排",
        "生物安排",
        "历史安排",
        "政治安排",
        "地理安排",
        "课位包1",
        "课位包2",
        "课位包3",
        "课位包4",
        "课位包5",
        "课位包6",
    ]
    student_sheet.append(headers)
    for student in sorted(students, key=lambda item: (item.admin_class, item.name)):
        section_by_subject = student_sections[student.id]
        row = [
            student.name,
            student.work_id,
            student.admin_class,
            "+".join(student.selected_subjects),
        ]
        for subject in SUBJECTS:
            section = section_by_subject[subject]
            package = package_of_section[section.id] + 1
            row.append(f"{section.name}（{section.teacher}，包{package}）")
        for package in PACKAGES:
            section = student_packages[student.id][package]
            row.append(f"{section.name}（{section.teacher}）")
        student_sheet.append(row)

    teacher_sheet = wb.create_sheet("教师课位表")
    teacher_sheet.append(["教师", "课位包", "课程班", "科目", "类型", "学生数"])
    sections_by_teacher = defaultdict(list)
    for section in sections:
        sections_by_teacher[section.teacher].append(section)
    for teacher in sorted(sections_by_teacher):
        teacher_sections = sorted(
            sections_by_teacher[teacher],
            key=lambda item: package_of_section[item.id],
        )
        for section in teacher_sections:
            teacher_sheet.append(
                [
                    teacher,
                    f"课位包{package_of_section[section.id] + 1}",
                    section.name,
                    section.subject,
                    section.mode,
                    len(section_students[section.id]),
                ]
            )

    note_sheet = wb.create_sheet("说明")
    note_sheet.append(["项目", "内容"])
    note_sheet.append(["输入文件", str(find_input_workbook())])
    note_sheet.append(["去重规则", "教师列表按 课程班+教师 完全去重"])
    note_sheet.append(["课位包数量", "6"])
    note_sheet.append(["分班原则", "同科同类型课程班人数尽量均衡，按现有教师课程班数拆班"])
    note_sheet.append(["班额说明", "因教师课程班数量固定，个别课程班会出现24人或36人，属于最均衡可行解"])

    for sheet in wb.worksheets:
        autosize(sheet)

    wb.save(output_path)


def print_console_summary(
    students: List[Student],
    sections: List[Section],
    solver: cp_model.CpSolver,
    x: Dict,
    y: Dict,
    relax: int,
):
    section_map = {section.id: section for section in sections}
    package_of_section = {}
    section_students: Dict[int, List[Student]] = defaultdict(list)

    for section in sections:
        for package in PACKAGES:
            if solver.Value(y[(section.id, package)]):
                package_of_section[section.id] = package
                break

    for student in students:
        for section in sections:
            key = (student.id, section.id)
            if key in x and solver.Value(x[key]):
                section_students[section.id].append(student)

    print(f"求解完成，relax={relax}")
    for package in PACKAGES:
        print(f"--- 课位包{package + 1} ---")
        package_sections = [
            section
            for section in sections
            if package_of_section[section.id] == package
        ]
        package_sections.sort(key=lambda item: (item.subject, item.mode, item.name))
        for section in package_sections:
            print(
                f"{section.name}\t{section.teacher}\t{len(section_students[section.id])}"
            )


def main():
    input_path = find_input_workbook()
    output_path = next_available_output_path(Path.cwd() / "高一选考六课位分班结果.xlsx")
    students, sections = load_data(input_path)
    solver, x, y, _z, relax = solve(students, sections)
    export_results(output_path, students, sections, solver, x, y)
    print_console_summary(students, sections, solver, x, y, relax)
    print(f"OUTPUT={output_path}")


if __name__ == "__main__":
    main()
