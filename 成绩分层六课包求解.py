from __future__ import annotations

import os
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook, load_workbook
from ortools.sat.python import cp_model


SUBJECTS = ["物理", "化学", "地理", "生物", "历史", "政治"]
SELECTED_SUBJECTS = ["物理", "化学", "生物", "历史", "地理", "政治"]
PACKAGES = list(range(6))
SUBJECT_PRIORITY = {
    "物理": 100_000,
    "化学": 70_000,
    "地理": 40_000,
    "生物": 25_000,
    "历史": 8_000,
    "政治": 8_000,
}
MODE_MULTIPLIER = {"选考": 2, "合格考": 1}
A_ENTRY_WEIGHT = {"物理": 35_000, "化学": 25_000}
A_SIZE_WEIGHT = {"物理": 300, "化学": 200}
SHORTFALL_WEIGHT = 10_000
BALANCE_WEIGHT = 20
TOTAL_CHANGE_WEIGHT = 1_000_000_000


def normalize_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return (
        text.replace("１", "1")
        .replace("２", "2")
        .replace("３", "3")
        .replace("Ａ", "A")
        .replace("Ｂ", "B")
        .replace("Ｃ", "C")
        .replace("，", ",")
        .replace(" ", "")
        .replace("合考", "合格考")
    )


def latest_input_workbook() -> Path:
    desktop = Path.home() / "Desktop"
    matches = [
        desktop / name
        for name in os.listdir(desktop)
        if name.endswith(".xlsx")
        and not name.startswith("~$")
        and "成绩分层" in name
    ]
    if not matches:
        raise FileNotFoundError("未找到成绩分层的源数据 Excel")
    matches.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return matches[0]


def next_available_output_path(base_path: Path) -> Path:
    if not base_path.exists():
        return base_path
    stem = base_path.stem
    suffix = base_path.suffix
    parent = base_path.parent
    index = 2
    while True:
        candidate = parent / f"{stem}_{index}{suffix}"
        if not candidate.exists():
            return candidate
        index += 1


def course_level(mode: str, text: str) -> Optional[str]:
    if mode == "选考":
        match = re.search(r"选考([ABC])", text)
    else:
        match = re.search(r"(?:合格考)([123])", text)
    return match.group(1) if match else None


def detect_subject(text: str) -> Optional[str]:
    return next((subject for subject in SUBJECTS if text.startswith(subject)), None)


@dataclass(frozen=True)
class SectionKey:
    subject: str
    mode: str
    level: str

    @property
    def label(self) -> str:
        return f"{self.subject}{self.mode}{self.level}"


@dataclass
class Section:
    id: int
    key: SectionKey
    display_name: str
    teacher: str


@dataclass
class OriginalPlacement:
    raw: str
    key: Optional[SectionKey]
    valid: bool
    reason: str = ""


@dataclass
class StudentSubjectNeed:
    subject: str
    mode: str
    original: OriginalPlacement


@dataclass
class Student:
    id: int
    name: str
    work_id: str
    admin_class: str
    combo: str
    needs: Dict[str, StudentSubjectNeed]


def parse_section_key(text: str, mode_hint: Optional[str] = None) -> Optional[SectionKey]:
    normalized = normalize_text(text)
    if not normalized:
        return None
    subject = detect_subject(normalized)
    if subject is None:
        return None
    if "选考" in normalized:
        mode = "选考"
    elif "合格考" in normalized:
        mode = "合格考"
    else:
        mode = mode_hint
    if mode is None:
        return None
    level = course_level(mode, normalized)
    if level is None:
        return None
    return SectionKey(subject=subject, mode=mode, level=level)


def parse_original_placement(raw: str, subject: str, mode: str) -> OriginalPlacement:
    normalized = normalize_text(raw)
    if not normalized:
        return OriginalPlacement(raw="", key=None, valid=False, reason="原始分层为空")
    key = parse_section_key(normalized, mode_hint=mode)
    if key is None:
        return OriginalPlacement(
            raw=normalized,
            key=None,
            valid=False,
            reason="原始分层命名异常，无法识别层次",
        )
    if key.subject != subject or key.mode != mode:
        return OriginalPlacement(
            raw=normalized,
            key=None,
            valid=False,
            reason="原始分层与科目/类型不一致",
        )
    return OriginalPlacement(raw=normalized, key=key, valid=True)


def normalized_display_name(text: str, key: SectionKey) -> str:
    normalized = normalize_text(text)
    if normalized.endswith("班"):
        return normalized
    return f"{normalized}班" if key.mode == "合格考" or key.subject == "物理" else normalized


def load_data(workbook_path: Path) -> Tuple[List[Student], List[Section]]:
    wb = load_workbook(workbook_path, data_only=True)
    selected_sheet = wb["选考-源数据"]
    qualified_sheet = wb["合格考-源数据"]
    teacher_sheet = wb["教师列表"]

    selected_rows = list(selected_sheet.iter_rows(min_row=2, values_only=True))
    qualified_rows = list(qualified_sheet.iter_rows(min_row=2, values_only=True))
    if len(selected_rows) != len(qualified_rows):
        raise ValueError("选考与合格考学生行数不一致")

    students: List[Student] = []
    selected_subject_columns = {
        "物理": 4,
        "化学": 5,
        "生物": 6,
        "历史": 7,
        "地理": 8,
        "政治": 9,
    }
    qualified_subject_columns = {
        "物理": 3,
        "化学": 4,
        "生物": 5,
        "历史": 6,
        "地理": 7,
        "政治": 8,
    }

    for idx, (selected_row, qualified_row) in enumerate(zip(selected_rows, qualified_rows)):
        name1, work_id1, admin1, combo = selected_row[:4]
        name2, work_id2, admin2 = qualified_row[:3]
        identity1 = (
            normalize_text(name1),
            normalize_text(work_id1),
            normalize_text(admin1),
        )
        identity2 = (
            normalize_text(name2),
            normalize_text(work_id2),
            normalize_text(admin2),
        )
        if identity1 != identity2:
            raise ValueError(f"第 {idx + 2} 行学生信息不一致：{identity1} vs {identity2}")

        combo_text = normalize_text(combo)
        chosen_subjects = {item for item in combo_text.split(",") if item}
        needs: Dict[str, StudentSubjectNeed] = {}
        for subject in SUBJECTS:
            selected_raw = selected_row[selected_subject_columns[subject]]
            qualified_raw = qualified_row[qualified_subject_columns[subject]]
            if subject in chosen_subjects:
                mode = "选考"
                if selected_raw:
                    original = parse_original_placement(selected_raw, subject, mode)
                else:
                    extra = normalize_text(qualified_raw)
                    reason = "原始选考分层缺失"
                    if extra:
                        reason += f"，合格考表存在异常值：{extra}"
                    original = OriginalPlacement(
                        raw=normalize_text(selected_raw),
                        key=None,
                        valid=False,
                        reason=reason,
                    )
            elif qualified_raw:
                mode = "合格考"
                original = parse_original_placement(qualified_raw, subject, mode)
            else:
                mode = "合格考"
                extra = normalize_text(selected_raw)
                reason = "原始合格考分层缺失"
                if extra:
                    reason += f"，选考表存在异常值：{extra}"
                original = OriginalPlacement(
                    raw=normalize_text(qualified_raw),
                    key=None,
                    valid=False,
                    reason=reason,
                )
            needs[subject] = StudentSubjectNeed(subject=subject, mode=mode, original=original)

        students.append(
            Student(
                id=idx,
                name=identity1[0],
                work_id=identity1[1],
                admin_class=identity1[2],
                combo=combo_text,
                needs=needs,
            )
        )

    seen = set()
    sections: List[Section] = []
    for row in teacher_sheet.iter_rows(min_row=2, values_only=True):
        course_raw, teacher_raw = row[:2]
        if not course_raw or not teacher_raw:
            continue
        teacher = normalize_text(teacher_raw)
        key = parse_section_key(course_raw)
        if key is None:
            continue
        display_name = normalized_display_name(str(course_raw), key)
        dedup_key = (key.subject, key.mode, key.level, teacher)
        if dedup_key in seen:
            continue
        seen.add(dedup_key)
        sections.append(
            Section(
                id=len(sections),
                key=key,
                display_name=display_name,
                teacher=teacher,
            )
        )

    return students, sections


def build_model(
    students: List[Student],
    sections: List[Section],
):
    model = cp_model.CpModel()

    sections_by_subject_mode: Dict[Tuple[str, str], List[Section]] = defaultdict(list)
    sections_by_key: Dict[SectionKey, Section] = {}
    sections_by_teacher: Dict[str, List[Section]] = defaultdict(list)
    for section in sections:
        sections_by_subject_mode[(section.key.subject, section.key.mode)].append(section)
        sections_by_key[section.key] = section
        sections_by_teacher[section.teacher].append(section)

    x = {}
    keep_original_terms = []
    invalid_originals = []
    change_vars = []
    weighted_change_terms = []
    balance_terms = []
    a_control_terms = []

    for student in students:
        for subject in SUBJECTS:
            need = student.needs[subject]
            candidates = sections_by_subject_mode[(subject, need.mode)]
            vars_for_subject = []
            for section in candidates:
                var = model.NewBoolVar(f"x_s{student.id}_{subject}_c{section.id}")
                x[(student.id, subject, section.id)] = var
                vars_for_subject.append(var)
            model.Add(sum(vars_for_subject) == 1)

            if need.original.valid and need.original.key in sections_by_key:
                original_section = sections_by_key[need.original.key]
                keep_var = x[(student.id, subject, original_section.id)]
                weight = SUBJECT_PRIORITY[subject] * MODE_MULTIPLIER[need.mode]
                change_var = model.NewBoolVar(f"changed_s{student.id}_{subject}")
                model.Add(change_var + keep_var == 1)
                change_vars.append(change_var)
                weighted_change_terms.append(weight * change_var)
                keep_original_terms.append((student, subject, keep_var, original_section))
            else:
                invalid_originals.append((student, subject))

    y = {}
    for section in sections:
        package_vars = []
        for package in PACKAGES:
            var = model.NewBoolVar(f"y_c{section.id}_p{package}")
            y[(section.id, package)] = var
            package_vars.append(var)
        model.Add(sum(package_vars) == 1)

    for package in PACKAGES:
        model.Add(sum(y[(section.id, package)] for section in sections) == 5)

    for teacher, teacher_sections in sections_by_teacher.items():
        for package in PACKAGES:
            model.Add(sum(y[(section.id, package)] for section in teacher_sections) <= 1)

    if sections:
        model.Add(y[(sections[0].id, 0)] == 1)

    z = {}
    for student in students:
        eligible_sections = []
        for subject in SUBJECTS:
            need = student.needs[subject]
            eligible_sections.extend(sections_by_subject_mode[(subject, need.mode)])
        for section in eligible_sections:
            for package in PACKAGES:
                z_var = model.NewBoolVar(f"z_s{student.id}_c{section.id}_p{package}")
                x_var = x[(student.id, section.key.subject, section.id)]
                y_var = y[(section.id, package)]
                model.Add(z_var <= x_var)
                model.Add(z_var <= y_var)
                model.Add(z_var >= x_var + y_var - 1)
                z[(student.id, section.id, package)] = z_var
        for package in PACKAGES:
            model.Add(
                sum(z[(student.id, section.id, package)] for section in eligible_sections) == 1
            )

    size_vars = {}
    shortfall_vars = {}
    deviation_vars = {}
    objective_meta = {}
    section_by_level_subject_mode = defaultdict(dict)
    for section in sections:
        section_by_level_subject_mode[(section.key.subject, section.key.mode)][section.key.level] = section

    totals_by_type = Counter()
    for student in students:
        for subject in SUBJECTS:
            need = student.needs[subject]
            totals_by_type[(subject, need.mode)] += 1

    for (subject, mode), type_sections in sections_by_subject_mode.items():
        total = totals_by_type[(subject, mode)]
        count = len(type_sections)
        average_floor = total // count
        feasible_low = 25 if total >= 25 * count else average_floor
        for section in type_sections:
            size_var = model.NewIntVar(feasible_low, 38, f"size_c{section.id}")
            model.Add(
                size_var
                == sum(
                    x[(student.id, subject, section.id)]
                    for student in students
                    if student.needs[subject].mode == mode
                )
            )
            size_vars[section.id] = size_var

            shortfall = model.NewIntVar(0, 25, f"shortfall_c{section.id}")
            model.AddMaxEquality(shortfall, [0, 25 - size_var])
            shortfall_vars[section.id] = shortfall
            balance_terms.append(SHORTFALL_WEIGHT * shortfall)

            scaled_deviation = model.NewIntVar(0, total * count, f"dev_c{section.id}")
            model.AddAbsEquality(scaled_deviation, size_var * count - total)
            deviation_vars[section.id] = scaled_deviation
            balance_terms.append(BALANCE_WEIGHT * scaled_deviation)

        # B/C composition constraint on A/B/C subjects.
        if mode == "选考" and set(section.key.level for section in type_sections) >= {"A", "B", "C"}:
            b_section = section_by_level_subject_mode[(subject, mode)]["B"]
            c_section = section_by_level_subject_mode[(subject, mode)]["C"]
            b_c_from_c = sum(
                x[(student.id, subject, b_section.id)]
                for student in students
                if student.needs[subject].mode == mode
                and student.needs[subject].original.valid
                and student.needs[subject].original.key
                and student.needs[subject].original.key.level == "C"
            )
            c_b_from_b = sum(
                x[(student.id, subject, c_section.id)]
                for student in students
                if student.needs[subject].mode == mode
                and student.needs[subject].original.valid
                and student.needs[subject].original.key
                and student.needs[subject].original.key.level == "B"
            )
            model.Add(2 * b_c_from_c <= size_vars[b_section.id])
            model.Add(2 * c_b_from_b <= size_vars[c_section.id])

    # Physics/Chemistry A classes: avoid extra students entering A and keep A smaller.
    for subject in ("物理", "化学"):
        a_section = section_by_level_subject_mode[(subject, "选考")]["A"]
        a_control_terms.append(A_SIZE_WEIGHT[subject] * size_vars[a_section.id])
        for student in students:
            need = student.needs[subject]
            if need.mode != "选考":
                continue
            original_level = need.original.key.level if need.original.valid and need.original.key else ""
            if original_level != "A":
                a_control_terms.append(
                    A_ENTRY_WEIGHT[subject] * x[(student.id, subject, a_section.id)]
                )

    for student, subject, keep_var, section in keep_original_terms:
        model.AddHint(keep_var, 1)
    for section in sections:
        model.AddHint(y[(section.id, section.id % 6)], 1)

    return (
        model,
        x,
        y,
        size_vars,
        shortfall_vars,
        deviation_vars,
        sections_by_key,
        change_vars,
        weighted_change_terms,
        balance_terms,
        a_control_terms,
    )


def solve(students: List[Student], sections: List[Section]):
    (
        model,
        x,
        y,
        size_vars,
        shortfall_vars,
        deviation_vars,
        sections_by_key,
        change_vars,
        weighted_change_terms,
        balance_terms,
        a_control_terms,
    ) = build_model(
        students, sections
    )
    total_changes = model.NewIntVar(0, len(change_vars), "total_changes")
    model.Add(total_changes == sum(change_vars))
    model.Minimize(total_changes)

    stage1_solver = cp_model.CpSolver()
    stage1_solver.parameters.max_time_in_seconds = 240
    stage1_solver.parameters.num_search_workers = 8
    stage1_solver.parameters.random_seed = 42
    stage1_solver.parameters.relative_gap_limit = 0.0
    stage1_status = stage1_solver.Solve(model)
    if stage1_status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"第一阶段没有找到可行解，状态={stage1_status}")

    best_total_changes = int(stage1_solver.Value(total_changes))
    model.Add(total_changes == best_total_changes)
    model.ClearObjective()
    model.Minimize(sum(weighted_change_terms) + sum(a_control_terms) + sum(balance_terms))
    model.ClearHints()
    for key, var in x.items():
        model.AddHint(var, stage1_solver.Value(var))
    for key, var in y.items():
        model.AddHint(var, stage1_solver.Value(var))

    solver = cp_model.CpSolver()
    solver.parameters.max_time_in_seconds = 480
    solver.parameters.num_search_workers = 8
    solver.parameters.random_seed = 42
    solver.parameters.relative_gap_limit = 0.001
    status = solver.Solve(model)
    if status not in (cp_model.OPTIMAL, cp_model.FEASIBLE):
        raise RuntimeError(f"第二阶段没有找到可行解，状态={status}")
    return (
        solver,
        x,
        y,
        size_vars,
        shortfall_vars,
        deviation_vars,
        sections_by_key,
        best_total_changes,
    )


def autosize(sheet):
    widths = defaultdict(int)
    for row in sheet.iter_rows():
        for cell in row:
            value = "" if cell.value is None else str(cell.value)
            widths[cell.column] = max(widths[cell.column], len(value))
    for column, width in widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = min(
            max(width + 2, 10), 60
        )


def build_solution_maps(
    students: List[Student],
    sections: List[Section],
    solver: cp_model.CpSolver,
    x: Dict,
    y: Dict,
):
    package_of_section = {}
    for section in sections:
        for package in PACKAGES:
            if solver.Value(y[(section.id, package)]):
                package_of_section[section.id] = package
                break

    section_students: Dict[int, List[Student]] = defaultdict(list)
    student_assignments: Dict[int, Dict[str, Section]] = defaultdict(dict)
    student_packages: Dict[int, Dict[int, Section]] = defaultdict(dict)
    for student in students:
        for subject in SUBJECTS:
            need = student.needs[subject]
            for section in sections:
                key = (student.id, subject, section.id)
                if key in x and solver.Value(x[key]):
                    section_students[section.id].append(student)
                    student_assignments[student.id][subject] = section
                    student_packages[student.id][package_of_section[section.id]] = section
                    break

    return package_of_section, section_students, student_assignments, student_packages


def original_level(student: Student, subject: str) -> str:
    need = student.needs[subject]
    return need.original.key.level if need.original.valid and need.original.key else ""


def reason_for_change(
    student: Student,
    subject: str,
    old_key: Optional[SectionKey],
    new_section: Section,
    package_of_section: Dict[int, int],
    sections_by_key: Dict[SectionKey, Section],
    student_assignments: Dict[int, Dict[str, Section]],
    section_sizes: Dict[int, int],
) -> str:
    if old_key is None:
        return student.needs[subject].original.reason or "原始分层缺失"
    if old_key not in sections_by_key:
        return "原始分层在教师班级表中不存在"
    old_section = sections_by_key[old_key]
    old_package = package_of_section[old_section.id]
    conflicting = []
    for other_subject, section in student_assignments[student.id].items():
        if other_subject == subject:
            continue
        if package_of_section[section.id] == old_package:
            conflicting.append(f"{other_subject}-{section.display_name}")
    if conflicting:
        return f"若保留原分层，将与 {', '.join(conflicting)} 发生课位冲突"
    if section_sizes[old_section.id] >= 38:
        return "原分层班额已达上限"
    return "为满足六课包整体可行、教师课位与班额均衡而调整"


def export_results(
    output_path: Path,
    students: List[Student],
    sections: List[Section],
    solver: cp_model.CpSolver,
    x: Dict,
    y: Dict,
    size_vars: Dict[int, cp_model.IntVar],
    shortfall_vars: Dict[int, cp_model.IntVar],
    sections_by_key: Dict[SectionKey, Section],
):
    package_of_section, section_students, student_assignments, student_packages = build_solution_maps(
        students, sections, solver, x, y
    )
    section_sizes = {section.id: solver.Value(size_vars[section.id]) for section in sections}

    wb = Workbook()
    wb.remove(wb.active)

    selected_package_sheet = wb.create_sheet("选科分层结果")
    selected_package_sheet.append(
        ["课位包", "课程班", "教师", "学生数", "原A人数", "原B人数", "原C人数", "学生名单"]
    )
    for package in PACKAGES:
        package_sections = [
            section
            for section in sections
            if package_of_section[section.id] == package and section.key.mode == "选考"
        ]
        package_sections.sort(key=lambda item: (item.key.subject, item.key.level))
        for section in package_sections:
            counts = Counter(
                original_level(student, section.key.subject)
                for student in section_students[section.id]
                if original_level(student, section.key.subject)
            )
            names = "、".join(
                student.name for student in sorted(section_students[section.id], key=lambda s: (s.admin_class, s.name))
            )
            selected_package_sheet.append(
                [
                    f"课位包{package + 1}",
                    section.display_name,
                    section.teacher,
                    section_sizes[section.id],
                    counts.get("A", 0),
                    counts.get("B", 0),
                    counts.get("C", 0),
                    names,
                ]
            )

    compare_sheet = wb.create_sheet("学生原始新分层对比")
    compare_headers = ["姓名", "学工号", "行政班", "新科目组合"]
    for subject in SUBJECTS:
        compare_headers.extend(
            [
                f"{subject}原始分层",
                f"{subject}新分层",
                f"{subject}是否变动",
                f"{subject}课位包",
            ]
        )
    compare_sheet.append(compare_headers)

    unsatisfied_sheet = wb.create_sheet("不能满足名单")
    unsatisfied_sheet.append(
        ["姓名", "学工号", "行政班", "科目", "类型", "原始分层", "新分层", "新课位包", "原因"]
    )

    stats_sheet = wb.create_sheet("分层变动统计")
    stats_sheet.append(
        ["科目", "类型", "层次", "原始人数", "新人数", "原层保留人数", "调出人数", "调入人数"]
    )

    package_summary_sheet = wb.create_sheet("每课位课程清单")
    package_summary_sheet.append(
        ["课位包", "课程班", "科目", "类型", "教师", "学生数", "学生名单"]
    )

    teacher_sheet = wb.create_sheet("教师课位表")
    teacher_sheet.append(["教师", "课位包", "课程班", "科目", "类型", "学生数"])

    level_stats: Dict[Tuple[str, str, str], Dict[str, int]] = defaultdict(lambda: defaultdict(int))

    for student in sorted(students, key=lambda item: (item.admin_class, item.name)):
        row = [student.name, student.work_id, student.admin_class, student.combo]
        for subject in SUBJECTS:
            need = student.needs[subject]
            new_section = student_assignments[student.id][subject]
            changed = not (
                need.original.valid
                and need.original.key == new_section.key
            )
            package_name = f"课位包{package_of_section[new_section.id] + 1}"
            original_text = need.original.raw or "无"
            row.extend(
                [
                    original_text,
                    new_section.display_name,
                    "是" if changed else "否",
                    package_name,
                ]
            )

            new_level = new_section.key.level
            if need.original.valid and need.original.key:
                old_level = need.original.key.level
                level_stats[(subject, need.mode, old_level)]["original"] += 1
                if old_level == new_level:
                    level_stats[(subject, need.mode, old_level)]["kept"] += 1
                else:
                    level_stats[(subject, need.mode, old_level)]["out"] += 1
                    level_stats[(subject, need.mode, new_level)]["in"] += 1
            elif need.original.raw:
                level_stats[(subject, need.mode, new_level)]["in"] += 1
            level_stats[(subject, need.mode, new_level)]["new"] += 1

            if changed:
                reason = reason_for_change(
                    student=student,
                    subject=subject,
                    old_key=need.original.key if need.original.valid else None,
                    new_section=new_section,
                    package_of_section=package_of_section,
                    sections_by_key=sections_by_key,
                    student_assignments=student_assignments,
                    section_sizes=section_sizes,
                )
                unsatisfied_sheet.append(
                    [
                        student.name,
                        student.work_id,
                        student.admin_class,
                        subject,
                        need.mode,
                        original_text,
                        new_section.display_name,
                        package_name,
                        reason,
                    ]
                )
        compare_sheet.append(row)

    sections_by_teacher = defaultdict(list)
    for section in sections:
        sections_by_teacher[section.teacher].append(section)
        students_in_section = sorted(
            section_students[section.id], key=lambda item: (item.admin_class, item.name)
        )
        package_summary_sheet.append(
            [
                f"课位包{package_of_section[section.id] + 1}",
                section.display_name,
                section.key.subject,
                section.key.mode,
                section.teacher,
                section_sizes[section.id],
                "、".join(student.name for student in students_in_section),
            ]
        )

    for teacher, teacher_sections in sorted(sections_by_teacher.items()):
        for section in sorted(teacher_sections, key=lambda item: package_of_section[item.id]):
            teacher_sheet.append(
                [
                    teacher,
                    f"课位包{package_of_section[section.id] + 1}",
                    section.display_name,
                    section.key.subject,
                    section.key.mode,
                    section_sizes[section.id],
                ]
            )

    levels_by_subject_mode = defaultdict(set)
    for section in sections:
        levels_by_subject_mode[(section.key.subject, section.key.mode)].add(section.key.level)

    for (subject, mode), levels in sorted(levels_by_subject_mode.items()):
        for level in sorted(levels):
            stats = level_stats[(subject, mode, level)]
            stats_sheet.append(
                [
                    subject,
                    mode,
                    level,
                    stats.get("original", 0),
                    stats.get("new", 0),
                    stats.get("kept", 0),
                    stats.get("out", 0),
                    stats.get("in", 0),
                ]
            )

    note_sheet = wb.create_sheet("说明")
    note_sheet.append(["项目", "内容"])
    note_sheet.append(["输入文件", str(latest_input_workbook())])
    note_sheet.append(["教师去重规则", "教师列表按 课程班层次+教师 完全去重"])
    note_sheet.append(["课位包数量", "6"])
    note_sheet.append(["求解方式", "OR-Tools CP-SAT 约束求解"])
    note_sheet.append(
        [
            "优化优先级",
            "物理>化学>地理>生物>历史=政治；优先保原成绩分层，再满足六课包、教师不冲突、班额均衡；物理/化学A额外控制少进人",
        ]
    )
    note_sheet.append(
        [
            "已知异常",
            "王梓萱(16117484) 合格考生物原始分层仅写“生物”，已按原始数据异常处理，并在不能满足名单中单列",
        ]
    )
    for section in sections:
        shortfall = solver.Value(shortfall_vars[section.id])
        if shortfall > 0:
            note_sheet.append(
                [
                    f"{section.display_name}班额提示",
                    f"该班 {section_sizes[section.id]} 人，低于 25 人 {shortfall} 人；受该科教师班数与总人数限制，已是可行最优范围内的结果",
                ]
            )

    for sheet in wb.worksheets:
        autosize(sheet)

    wb.save(output_path)


def export_teacher_brief(
    output_path: Path,
    students: List[Student],
    sections: List[Section],
    solver: cp_model.CpSolver,
    x: Dict,
    y: Dict,
    size_vars: Dict[int, cp_model.IntVar],
):
    package_of_section, section_students, student_assignments, student_packages = build_solution_maps(
        students, sections, solver, x, y
    )
    section_sizes = {section.id: solver.Value(size_vars[section.id]) for section in sections}

    wb = Workbook()
    wb.remove(wb.active)

    summary_sheet = wb.create_sheet("课位包总表")
    summary_sheet.append(["课位包", "课程班", "科目", "类型", "教师", "学生数"])
    for package in PACKAGES:
        package_sections = [section for section in sections if package_of_section[section.id] == package]
        package_sections.sort(key=lambda item: (item.key.mode, item.key.subject, item.key.level))
        for section in package_sections:
            summary_sheet.append(
                [
                    f"课位包{package + 1}",
                    section.display_name,
                    section.key.subject,
                    section.key.mode,
                    section.teacher,
                    section_sizes[section.id],
                ]
            )

    student_sheet = wb.create_sheet("学生课位一览")
    student_sheet.append(
        ["姓名", "学工号", "行政班", "选科组合", "课位包1", "课位包2", "课位包3", "课位包4", "课位包5", "课位包6"]
    )
    for student in sorted(students, key=lambda item: (item.admin_class, item.name)):
        row = [student.name, student.work_id, student.admin_class, student.combo]
        for package in PACKAGES:
            section = student_packages[student.id][package]
            row.append(f"{section.display_name}（{section.teacher}）")
        student_sheet.append(row)

    roster_sheet = wb.create_sheet("课程班名单")
    roster_sheet.append(["课位包", "课程班", "教师", "学生数", "学生名单"])
    for package in PACKAGES:
        package_sections = [section for section in sections if package_of_section[section.id] == package]
        package_sections.sort(key=lambda item: (item.key.mode, item.key.subject, item.key.level))
        for section in package_sections:
            names = "、".join(
                student.name
                for student in sorted(section_students[section.id], key=lambda item: (item.admin_class, item.name))
            )
            roster_sheet.append(
                [
                    f"课位包{package + 1}",
                    section.display_name,
                    section.teacher,
                    section_sizes[section.id],
                    names,
                ]
            )

    teacher_sheet = wb.create_sheet("教师课位一览")
    teacher_sheet.append(["教师", "课位包", "课程班", "学生数"])
    sections_by_teacher = defaultdict(list)
    for section in sections:
        sections_by_teacher[section.teacher].append(section)
    for teacher in sorted(sections_by_teacher):
        for section in sorted(sections_by_teacher[teacher], key=lambda item: package_of_section[item.id]):
            teacher_sheet.append(
                [
                    teacher,
                    f"课位包{package_of_section[section.id] + 1}",
                    section.display_name,
                    section_sizes[section.id],
                ]
            )

    for sheet in wb.worksheets:
        autosize(sheet)
    wb.save(output_path)


def print_summary(
    students: List[Student],
    sections: List[Section],
    solver: cp_model.CpSolver,
    x: Dict,
    y: Dict,
    size_vars: Dict[int, cp_model.IntVar],
):
    package_of_section, section_students, _, _ = build_solution_maps(students, sections, solver, x, y)
    print("求解完成")
    for package in PACKAGES:
        print(f"--- 课位包{package + 1} ---")
        package_sections = [
            section for section in sections if package_of_section[section.id] == package
        ]
        package_sections.sort(key=lambda item: (item.key.mode, item.key.subject, item.key.level))
        for section in package_sections:
            print(
                f"{section.display_name}\t{section.teacher}\t{solver.Value(size_vars[section.id])}"
            )


def main():
    input_path = latest_input_workbook()
    output_path = next_available_output_path(Path.cwd() / "高一成绩分层六课包结果.xlsx")
    teacher_brief_path = next_available_output_path(Path.cwd() / "高一成绩分层六课包结果_教师简洁版.xlsx")
    students, sections = load_data(input_path)
    (
        solver,
        x,
        y,
        size_vars,
        shortfall_vars,
        _deviation_vars,
        sections_by_key,
        best_total_changes,
    ) = solve(students, sections)
    export_results(
        output_path=output_path,
        students=students,
        sections=sections,
        solver=solver,
        x=x,
        y=y,
        size_vars=size_vars,
        shortfall_vars=shortfall_vars,
        sections_by_key=sections_by_key,
    )
    export_teacher_brief(
        output_path=teacher_brief_path,
        students=students,
        sections=sections,
        solver=solver,
        x=x,
        y=y,
        size_vars=size_vars,
    )
    print_summary(students, sections, solver, x, y, size_vars)
    print(f"TOTAL_CHANGES={best_total_changes}")
    print(f"OUTPUT={output_path}")
    print(f"TEACHER_OUTPUT={teacher_brief_path}")


if __name__ == "__main__":
    main()
